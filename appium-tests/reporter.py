import os
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class AppiumExcelReporter:
    def __init__(self, output_path="test_report.xlsx"):
        self.output_path = output_path
        self.steps = []
        self.start_time = datetime.datetime.now()
        
    def add_step(self, step_name, status, duration_ms, details="", screenshot_path=None):
        """
        Record a test step result.
        :param step_name: Name of the executed action/step
        :param status: 'PASS' or 'FAIL'
        :param duration_ms: Step execution duration in milliseconds
        :param details: Additional log context or error message
        :param screenshot_path: Path to captured screenshot on device
        """
        self.steps.append({
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "step_name": step_name,
            "status": status.upper(),
            "duration_ms": duration_ms,
            "details": details,
            "screenshot": screenshot_path or ""
        })

    def generate_report(self):
        """
        Compile and write results into a professional styled Excel spreadsheet.
        """
        wb = Workbook()
        
        # 1. SETUP SHEETS
        ws_summary = wb.active
        ws_summary.title = "Summary Dashboard"
        ws_details = wb.create_sheet(title="Test Details")
        
        # Enable grid lines visibility
        ws_summary.views.sheetView[0].showGridLines = True
        ws_details.views.sheetView[0].showGridLines = True
        
        # 2. DEFINE PALETTE & STYLES (Professional Teal theme)
        font_family = "Segoe UI"
        
        # Fonts
        font_title = Font(name=font_family, size=16, bold=True, color="1F4E79")
        font_section = Font(name=font_family, size=12, bold=True, color="2C3E50")
        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_bold = Font(name=font_family, size=10, bold=True)
        font_regular = Font(name=font_family, size=10)
        font_details = Font(name=font_family, size=9, italic=True, color="555555")
        
        # Fills
        fill_header = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")  # Deep Teal
        fill_zebra = PatternFill(start_color="F2F7FA", end_color="F2F7FA", fill_type="solid")   # Very light blue-gray
        fill_pass = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")    # Soft green
        fill_fail = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")    # Soft red
        fill_accent = PatternFill(start_color="E9ECEF", end_color="E9ECEF", fill_type="solid")  # Neutral accent
        
        # Borders
        thin_border_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
        
        thick_bottom_side = Side(border_style="medium", color="1F4E79")
        border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
        
        # Alignments
        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        
        # 3. BUILD SUMMARY DASHBOARD
        end_time = datetime.datetime.now()
        total_time = end_time - self.start_time
        total_steps = len(self.steps)
        passed_steps = sum(1 for s in self.steps if s["status"] == "PASS")
        failed_steps = sum(1 for s in self.steps if s["status"] == "FAIL")
        pass_rate = (passed_steps / total_steps * 100) if total_steps > 0 else 0.0
        
        # Sheet Title
        ws_summary.merge_cells("A1:D1")
        ws_summary["A1"] = "SkillSphereAI Mobile E2E Test Report"
        ws_summary["A1"].font = font_title
        ws_summary["A1"].alignment = align_left
        ws_summary.row_dimensions[1].height = 40
        
        # Divider Line
        ws_summary.row_dimensions[2].height = 10
        
        # Meta info
        ws_summary["A3"] = "Test Summary & Performance Indicators"
        ws_summary["A3"].font = font_section
        ws_summary.row_dimensions[3].height = 25
        
        metrics = [
            ("Execution Start", self.start_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("Execution End", end_time.strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Duration", f"{total_time.total_seconds():.2f} seconds"),
            ("Total Test Steps", total_steps),
            ("Steps Passed", passed_steps),
            ("Steps Failed", failed_steps),
            ("Pass Rate", f"{pass_rate:.1f}%")
        ]
        
        current_row = 4
        for metric_name, value in metrics:
            ws_summary.cell(row=current_row, column=1, value=metric_name).font = font_bold
            ws_summary.cell(row=current_row, column=1).alignment = align_left
            ws_summary.cell(row=current_row, column=1).fill = fill_accent
            ws_summary.cell(row=current_row, column=1).border = border_all
            
            val_cell = ws_summary.cell(row=current_row, column=2, value=value)
            val_cell.font = font_regular
            val_cell.alignment = align_left
            val_cell.border = border_all
            
            # Format Pass Rate separately
            if metric_name == "Pass Rate":
                val_cell.font = Font(name=font_family, size=10, bold=True, color="2E7D32" if pass_rate >= 80 else "C62828")
                val_cell.fill = fill_pass if pass_rate >= 80 else fill_fail
            elif metric_name == "Steps Failed" and failed_steps > 0:
                val_cell.font = Font(name=font_family, size=10, bold=True, color="C62828")
                val_cell.fill = fill_fail
            elif metric_name == "Steps Passed" and passed_steps > 0:
                val_cell.font = Font(name=font_family, size=10, bold=True, color="2E7D32")
                
            ws_summary.row_dimensions[current_row].height = 20
            current_row += 1
            
        # 4. BUILD DETAILS SHEET
        details_headers = ["No.", "Timestamp", "Step Name", "Status", "Duration (ms)", "Execution Details", "Screenshot Path"]
        
        # Header Row
        for col_idx, header in enumerate(details_headers, 1):
            cell = ws_details.cell(row=1, column=col_idx, value=header)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = border_header
            
        ws_details.row_dimensions[1].height = 26
        
        # Data Rows
        for row_idx, step in enumerate(self.steps, 2):
            ws_details.cell(row=row_idx, column=1, value=row_idx - 1).alignment = align_center
            ws_details.cell(row=row_idx, column=2, value=step["timestamp"]).alignment = align_center
            ws_details.cell(row=row_idx, column=3, value=step["step_name"]).alignment = align_left
            
            status_cell = ws_details.cell(row=row_idx, column=4, value=step["status"])
            status_cell.alignment = align_center
            status_cell.font = font_bold
            if step["status"] == "PASS":
                status_cell.fill = fill_pass
                status_cell.font = Font(name=font_family, size=10, bold=True, color="155724")
            else:
                status_cell.fill = fill_fail
                status_cell.font = Font(name=font_family, size=10, bold=True, color="721C24")
                
            ws_details.cell(row=row_idx, column=5, value=step["duration_ms"]).alignment = align_right
            
            details_cell = ws_details.cell(row=row_idx, column=6, value=step["details"])
            details_cell.alignment = align_left
            if step["status"] == "FAIL":
                details_cell.font = Font(name=font_family, size=9, bold=True, color="721C24")
            else:
                details_cell.font = font_details
                
            screenshot_cell = ws_details.cell(row=row_idx, column=7, value=step["screenshot"])
            screenshot_cell.alignment = align_left
            screenshot_cell.font = font_details
            
            # Apply common borders and alternate background
            for col_idx in range(1, 8):
                c = ws_details.cell(row=row_idx, column=col_idx)
                c.border = border_all
                if row_idx % 2 == 1 and step["status"] != "PASS" and step["status"] != "FAIL":
                    c.fill = fill_zebra
                    
            ws_details.row_dimensions[row_idx].height = 20
            
        # 5. AUTO-FIT COLUMN WIDTHS FOR BOTH SHEETS
        for ws in [ws_summary, ws_details]:
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        # Avoid checking merged cell length in a bad way
                        val_str = str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                # Pad and bound the column width
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        # Make specific summary columns extra readable
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 30
        ws_details.column_dimensions['F'].width = 50
        ws_details.column_dimensions['G'].width = 40
        
        # Save Workbook
        wb.save(self.output_path)
        print(f"Excel test analysis report saved successfully to: {os.path.abspath(self.output_path)}")


if __name__ == "__main__":
    # Self-test/verification: Let's run and verify a dummy report can be created
    reporter = AppiumExcelReporter("self_test_report.xlsx")
    reporter.add_step("Launch App", "PASS", 850, "App launched and main page rendered.")
    reporter.add_step("Authentication Flow", "PASS", 1200, "Successfully logged in as Tony Candidate.")
    reporter.add_step("Academics Details Fill", "PASS", 950, "Completed college credentials form step.")
    reporter.add_step("Gemini AI Resume Analysis", "PASS", 2400, "Gemini parser finished analysis. Score 85 returned.")
    reporter.add_step("Interactive Coding Assessment", "FAIL", 1800, "Assertion Error: Score did not increase by 10 points after correct code solution submission.", "screenshots/fail_step_5.png")
    reporter.generate_report()
